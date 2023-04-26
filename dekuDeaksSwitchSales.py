import time
import requests
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
import re
import csv
import locale
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import numbers
from openpyxl import load_workbook
import datetime

# DekuDeals URL #
countryParseList = []
userChoices = []
userInput = ""

#countryParse = {[countryNameList]: [currencySymbolList, currencyCodeList, countryButtonList]}
countryParse = {}

#date, time, currency declarations
today = datetime.date.today()
date_str = today.strftime("%d_%m_%Y")
#locale.setlocale(locale.LC_MONETARY, 'en_US.UTF-8')
#fileName = f"Nintendo Switch Game Discounts {date_str}.csv"


###################### Workig Code to write to excel ##############################

### go out and parse the country list to update the countryParse dictionary ###
edge_path = "D:\\Documents\\Programming\\msedgedriver.exe"
driver = webdriver.Edge(edge_path)
wait = WebDriverWait(driver, 15)
urls = ["https://www.dekudeals.com/hottest"]

### Opens and navigates to the USD sales section ###
page = driver.get(urls[0])
driver.maximize_window()
html = driver.page_source
soup = BeautifulSoup(html, "html.parser")

# Cookie disclaimer away
cookie_accepted = False
if not cookie_accepted:
    try:
        cookie_accept_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[class='fc-button fc-cta-consent fc-primary-button']")))
        cookie_accept_button.click()
        cookie_accepted = True
    except NoSuchElementException:
        pass

countrySelectorElement = driver.find_element(By.ID, "navbarCountry1")
countrySelectorElement.click()
dropdown = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='dropdown-menu country-select show']")))
countries = dropdown.find_elements(By.XPATH, "//button[@class='dropdown-item']")
countryButtonListDuplicates = []

countryNameList = []
currencySymbolList = []
currencyCodeList = []
countryButtonList = []
for country in countries:

    #retrieves the button element name
    span = country.find_element(By.XPATH, ".//span")
    countryButtonNameDuplicates = span.get_attribute("class")
    countryButtonListDuplicates.append(countryButtonNameDuplicates)

    #Retreives the country, currency symbol, and currency abbreviation
    country = country.text.strip()
    countryParseList.append(country)
    #removes all white lists
    countryParseList = [x for x in countryParseList if x != '']

    #strips the country name into multiple variables
    if country != "":
        countryName = re.search(r"^[A-Za-z\s]+", country).group().strip()
        #print(countryName)
        countryNameList.append(countryName)
        #strips the currency symbol and currency code
        currency = re.findall(r"\((.+),\s*(\w+)\)", country)[0]
        #strips the currncy symbol into a variable
        currencySymbol = currency[0].strip()
        #print(currencySymbol)
        currencySymbolList.append(currencySymbol)
        #print(currencySymbol)
        #strips the currncy code into a variable
        currencyCode = currency[1].strip()
        #print(currencyCode)
        currencyCodeList.append(currencyCode)
        #print(currencyCode)
    else:
        pass

#removes the duplicate button element names and creates a new list
countryButtonListUnformatted = []
for button in countryButtonListDuplicates:
    if button not in countryButtonListUnformatted:
        countryButtonListUnformatted.append(button)

#adds the proper syntax to the 
for button in countryButtonListUnformatted:
    if button not in countryButtonList:
        button = button.replace(" ", ".")
        countryButtonList.append("button[name='country'] > span." + button)

for i in range(len(countryNameList)):
    countryParse[countryNameList[i]] = [currencySymbolList[i], currencyCodeList[i], countryButtonList[i]]

driver.quit()


### Start of the programs user input ###
print("""
Hello and welcome to the Nintendo discount finder.
This unique app allows you to choose one or more region's Nintendo Switch sales.
Below are the region choices:

Country:         Currency:""")
for country, values in countryParse.items():
    print(f"""{country.ljust(15)} - {values[1].rjust(5)}""")

userInput = input("\n" + "Please enter the country name here: ").title()

while userInput not in countryParse:
    userInput = input("Not a valid entry sorry. Please check your spelling and choice: ").title()
userChoices.append(userInput)

chooseAnother = True
print("Would you like to choose another region?")
while chooseAnother == True:
    print(userChoices, "\n \n")
    userInput = input("Enter another region or hit the enter key to start: ").title()
    if userInput == "":
        chooseAnother = False
        break
    elif userInput in userChoices:
        print("\nCountry already selected. Entry not added.")
    elif userInput not in countryParse:
        print("\nNot a valid entry sorry. Please check your spelling and choice:")
    else:
        userChoices.append(userInput)
    
print("Thank you for your entry.")
print("We will start gathering the info you requested (" + str(userChoices) + ")")
print("Please wait")


##################### Parse start ##########################
edge_path = "D:\\Documents\\Programming\\msedgedriver.exe"
driver = webdriver.Edge(edge_path)
wait = WebDriverWait(driver, 15)
urls = ["https://www.dekudeals.com/hottest"]
cookie_accepted = False

for region in userChoices:
    print(countryParse[region][2])
    ### Opens and navigates to the USD sales section ###
    page = driver.get(urls[0])
    driver.maximize_window()

    html = driver.page_source
    soup = BeautifulSoup(html, "html.parser")

    # Cookie disclaimer away
    if not cookie_accepted:
        try:
            cookie_accept_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[class='fc-button fc-cta-consent fc-primary-button']")))
            cookie_accept_button.click()
            cookie_accepted = True
        except NoSuchElementException:
            pass


    countrySelector = soup.find("div", {"class": "dropdown-menu country-select"})
    countryCode = countrySelector.find_all((By.CLASS_NAME, "dropdown-item"))
    
    countrySelector = wait.until(EC.visibility_of_element_located((By.ID, "navbarCountry1")))
    countrySelector = driver.find_element(By.ID, "navbarCountry1")
    countrySelector.click()
    countrySelector = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, countryParse[region][2])))
    countrySelector.click()
    #filterSelector = driver.find_element(By.CSS_SELECTOR, "div.card-body")

    try:
        if wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "a[href*='?filter[store]=eshop']"))):
            print("eshop filter found")
            nintendoEshopFilter = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "a[href*='?filter[store]=eshop']")))
            #nintendoEshopFilter.click()
            ActionChains(driver).move_to_element(nintendoEshopFilter).click().perform()
            pass
        else:
            try:
                if wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "a[href*='?filter[format]=digital']"))):
                    print("digital filter found")
                    digitalGameFilter = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "a[href*='?filter[format]=digital']")))
                    #digitalGameFilter.click()
                    ActionChains(driver).move_to_element(digitalGameFilter).click().perform()
                    pass
            except Exception as e:
                print("No digital filter found after finding eshop filter")
                print(f"The error was {e}")
                pass
    except Exception as e:
        print("No eshop or digital filter found")
        print(f"The error was {e}")
        pass

    time.sleep(5)



    ## Stores the collected DekuDeals data into lists # 

    titleList = []
    normalPriceList = []
    discountPriceList = []
    discountPriceExpireList = []
    listIndex = 0

    while True:
        time.sleep(1)
        html = driver.page_source
        soup = BeautifulSoup(html, "html.parser")

        #search result container for individual game *subject to change per filter result url
        resultPage = soup.select('div[class*="col-xl-2 col-lg-3 col-sm-4 col-6 cell"]')

        try:
            #for result in resultPage:
            for result in resultPage:
                #title = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "a[class='main-link'] > div.h6 name")))
                title = result.select_one('a.main-link > div.h6.name').text
                #title = title.title()
                title = title.replace("\n", "")
                title = title.replace("'S", "'s")
                titleList.append(title)
                print(title)

                #price parse
                cardBadgeElement = result.find("div", {"class": "card-badge"})
                #normalPrice = result.find("div", {"class": "card-badge"}).text.split()[0]
                normalPrice = cardBadgeElement.find("s").text
                print("Normal price: " + normalPrice)
                normalPriceList.append(normalPrice)

                discountPrice = cardBadgeElement.find("strong").text
                #discountPrice = result.find("div", {"class": "card-badge"}).text.split()[1]
                discountPrice = discountPrice.replace("Regular Price:", "")
                discountPriceList.append(discountPrice)
                print("Current discounted price: " + discountPrice)

                ## Section to discern Availabity tag from Nintendo website ##
                #discountExpire is avialable
                if result.find("div", class_="w-100").find("small"):
                    discountExpire = result.find("div", class_="w-100").find("small").text
                    print("Found:", discountExpire + "\n")
                    discountPriceExpireList.append(discountExpire)
                #if theres a sale but no end date
                else:
                    print("(No discount date end provided)" + "\n")
                    discountPriceExpireList.append("-")

        except Exception as e:
                print(f"Error occurred for title: {title}")
                print(f"Error message: {e}" + "\n")
        
        #Navigates to next page of results until last page is reached
        pageNavigation = driver.find_element(By.CSS_SELECTOR, "div.pagination_controls")
        pagination = pageNavigation.find_element(By.CSS_SELECTOR, "ul.pagination")
        nextPageButton = pagination.find_elements(By.TAG_NAME, "li")[-1]
        if "disabled" in nextPageButton.get_attribute("class"):
            break
        nextPageButton.click()

    # Opens an excel workbook and creates the first row #
    #countryCurrencyFormat = "numbers.FORMAT_CURRENCY_" + countryParse[region][0]
    wb = Workbook()
    ws = wb.active
    ws.title = "Switch Discounts " + region
    ws.append(["Game Title", "Normal Price", "Discount Price", "Discount End"])
    header_font = Font(name="Calibri", size=12, bold=True, underline="single")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment

    for i in range(len(titleList)):
        ws.append([titleList[i], normalPriceList[i], discountPriceList[i], discountPriceExpireList[i]])
        ws.cell(row=i+2, column=2).number_format = "General"
        ws.cell(row=i+2, column=3).number_format = "General"
        listIndex += 1

    for col in range(2, 5):
        for row in range(2, len(titleList) + 2):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="right")

    wb.save(f"Switch discounts - {region} - {date_str}.xlsx")

####################### Workig Code to write to excel ##############################




###################### Workig Code to write to csv ##############################
#resultPage = soup.find_all("div", {"class": "BasicTilestyles__Container-sc-1bsju6x-0 krhhcK"})
#
#with open(fileName, mode="w", newline="") as file:
#    writer = csv.writer(file)
#    writer.writerow(["Game Title", "Normal Price", "Discount Price", "Discount Ends"])
#
#    for result in resultPage:
#        title = result.find("h3").text
#        normalPrice = float(result.find("span", {"class": "Pricestyles__MSRP-sc-1f0n8u6-5 ckvwhD"}).text.split("$")[1])
#        discountPrice = float(result.find("span", {"class": "Pricestyles__SalePrice-sc-1f0n8u6-4 dHJCdy"}).text.split("$")[1])
#        discountExpire = result.find("div", {"class": "ProductTilestyles__DescriptionTag-sc-m1loqs-5 ipvMnK"}).text.split(": ")[1]
#
#        normalPrice = locale.currency(normalPrice, grouping=True)
#        discountPrice = locale.currency(discountPrice, grouping=True)
#
#        writer.writerow([title, normalPrice, discountPrice, discountExpire])
###################### Workig Code to print to terminal ##############################



###################### Code to print to terminal ##############################

## Opens and navigates to the nintendo store search page #
#page = driver.get(urls[0])
#driver.maximize_window()
#time.sleep(2)
#salesFilter = driver.find_element(By.NAME, "Deals")
#salesFilter.click()
#time.sleep(5)
#
#html = driver.page_source
#soup = BeautifulSoup(html, "html.parser")
#
#titleList = []
#normalPriceList = []
#discountPriceList = []
#discountExpireList = []
#listIndex = 0
#
#resultPage = soup.find_all("div", {"class": "BasicTilestyles__Container-sc-1bsju6x-0 krhhcK"})
#
#for result in resultPage:
#    title = result.find("h3").text
#    titleList.append(title)
#    normalPrice = result.find("span", {"class": "Pricestyles__MSRP-sc-1f0n8u6-5 ckvwhD"}).text.split("$")[1]
#    normalPriceList.append(normalPrice)
#    price = result.find("span", {"class": "Pricestyles__SalePrice-sc-1f0n8u6-4 dHJCdy"}).text.split("$")[1]
#    discountPriceList.append(price)
#    discountExpire = result.find("div", {"class": "ProductTilestyles__DescriptionTag-sc-m1loqs-5 ipvMnK"}).text.split(": ")[1]
#    discountExpireList.append(discountExpire)
#
#
### Opens and navigates to the metacritic search page #
#page = driver.get(urls[1])
#driver.maximize_window()
#
## Cookie Button Away #
#time.sleep(5)
#cookieButton = driver.find_element(By.ID, "onetrust-accept-btn-handler")
#cookieButton.click()
#time.sleep(5)
#
#for title in titleList:
#    #search for the game
#    page = driver.get(urls[1])
#    search = driver.find_element(By.ID, "searchtextbox")
#    newString = re.sub(r"\s*\+\s*", " ", title).strip().lower()
#    search.send_keys(newString)
#    search.send_keys(Keys.ENTER)
#
#    #click game in search results
#    time.sleep(5)
#    searchResults = driver.find_elements(By.CLASS_NAME, "search_result")
#    for results in searchResults:
#        if "NS" in results.text:
#            print("Found NS")
#            name = results.find_element(By.CSS_SELECTOR, "div.sr_name a")
#            name.click()
#            break
#        elif " platform" in results.text:
#            print("Found Multiplatform")
#            name = results.find_element(By.CSS_SELECTOR, "div.sr_name a")
#            name.click()
#            break
#        else:
#            print("Could not find search result element")
#            pass
#    try:
#        #click review button
#        time.sleep(5)
#        reviewButton = driver.find_element(By.LINK_TEXT, 'Reviews')
#        reviewButton.click()
#        metaScore = None
#        userScore = None
#        platform = "NS"
#
#        #select nintendo switch if available
#        try:
#            time.sleep(5)
#            filterDropdown = None
#
#            try:
#                filterDropdown = driver.find_element(By.CSS_SELECTOR, 'select[onchange="mc_data_filter($(this).val());"]')
#            except NoSuchElementException:
#                print("No dropdown menu")
#                pass
#
#            if filterDropdown is not None and filterDropdown.is_displayed():
#                nsOption = None
#                for option in filterDropdown.find_elements(By.TAG_NAME, "option"):
#                    if "NS" in option.text:
#                       print("Dropdown is displayed")
#                       nsOption = option
#                       break
#
#                if nsOption is not None:
#                    print("Dropdown is displayed and 'NS' present")
#                    option.click()
#                    #gather metacritic scores after selecting Switch console
#                    time.sleep(5)
#                    html = driver.page_source
#                    soup = BeautifulSoup(html, "html.parser")
#                    selectNsScores = soup.find("div", {"class": "pod mcdata"})
#                    scoreRange = selectNsScores.find("ol", {"class": "list flex col2 ai_normal mc_module"})
#                    theScores = scoreRange.find_all('div', {'class': re.compile('rev_score large score_')})
#                    metaScore = theScores[0].text.strip()
#                    userScore = theScores[1].text.strip()
#                else:
#                    print("Dropdown is displayed and 'NS' is not present" + "\n")
#                    metaScore = "n/a"
#                    userScore = "n/a"
#                    
#            else:
#                print("Dropdown is not displayed" + "\n")
#                html = driver.page_source
#                soup = BeautifulSoup(html, "html.parser")
#                scoreRange = soup.find("ol", {"class": "list flex col2 ai_normal mc_module"})
#                theScores = scoreRange.find_all('div', {'class': re.compile('rev_score large score_')})
#                metaScore = theScores[0].text.strip()
#                userScore = theScores[1].text.strip()
#
#        except Exception as e: 
#            print(f"Error occurred for title: {title}")
#            print("Any other exception occured besides 'no such element'")
#            print(f"Error message: {e}" + "\n")
#            pass
#
#    except Exception as e:
#        #if review button is unavailable
#        print(f"Error occurred for title: {title}")
#        print("Review button not found")
#        print(f"Error message: {e}" + "\n")
#        metaScore = "n/a"
#        userScore = "n/a"
#        pass
#
#    print("Title: " + title.title())
#    print("Normal Price: $" + normalPriceList[listIndex])
#    print("Discount Price: $" + discountPriceList[listIndex])
#    print("Discount Expires: " + discountExpireList[listIndex])
#    print("Meta Score: " + metaScore)
#    print("User Score: " + userScore + "\n")
#    listIndex += 1
#
###################### Code to print to terminal ##############################




















